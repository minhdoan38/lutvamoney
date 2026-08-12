#!/usr/bin/env python3
"""Extract one table from a local data or document file.

Original utility maintained by Rylai for Codex and Hermes skill workflows.
"""

from __future__ import annotations

import argparse
import csv
from html.parser import HTMLParser
import json
import os
from pathlib import Path
import sys
import tempfile


class HtmlTableReader(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self.current_table: list[list[str]] | None = None
        self.current_row: list[str] | None = None
        self.current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag == "table" and self.current_table is None:
            self.current_table = []
        elif tag == "tr" and self.current_table is not None:
            self.current_row = []
        elif tag in {"td", "th"} and self.current_row is not None:
            self.current_cell = []

    def handle_data(self, data: str) -> None:
        if self.current_cell is not None:
            self.current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self.current_cell is not None and self.current_row is not None:
            value = " ".join("".join(self.current_cell).split())
            self.current_row.append(value)
            self.current_cell = None
        elif tag == "tr" and self.current_row is not None and self.current_table is not None:
            if self.current_row:
                self.current_table.append(self.current_row)
            self.current_row = None
        elif tag == "table" and self.current_table is not None:
            if self.current_table:
                self.tables.append(self.current_table)
            self.current_table = None


def require_file(path: Path) -> Path:
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"source file not found: {resolved}")
    return resolved


def normalize_rows(rows) -> list[list[str]]:
    normalized = []
    for row in rows:
        normalized.append(["" if value is None else str(value) for value in row])
    while normalized and not any(cell for cell in normalized[-1]):
        normalized.pop()
    return normalized


def load_json_tables(source: Path) -> list[list[list[str]]]:
    value = json.loads(source.read_text(encoding="utf-8-sig"))
    if isinstance(value, list) and all(isinstance(item, dict) for item in value):
        headers: list[str] = []
        for item in value:
            for key in item:
                key_text = str(key)
                if key_text not in headers:
                    headers.append(key_text)
        return [[headers] + [[str(item.get(key, "")) for key in headers] for item in value]]
    if isinstance(value, list) and all(isinstance(item, list) for item in value):
        return [normalize_rows(value)]
    raise RuntimeError("JSON table input must be a list of objects or a list of rows")


def load_tables(source: Path, sheet: str | None) -> list[list[list[str]]]:
    extension = source.suffix.lower()
    if extension in {".csv", ".tsv"}:
        delimiter = "\t" if extension == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return [normalize_rows(csv.reader(handle, delimiter=delimiter))]
    if extension == ".json":
        return load_json_tables(source)
    if extension in {".html", ".htm"}:
        reader = HtmlTableReader()
        reader.feed(source.read_text(encoding="utf-8-sig"))
        return reader.tables
    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError("install openpyxl to extract XLSX tables") from error
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            worksheets = [workbook[sheet]] if sheet else workbook.worksheets
            return [
                normalize_rows(worksheet.iter_rows(values_only=True))
                for worksheet in worksheets
            ]
        finally:
            workbook.close()
    if extension == ".pdf":
        try:
            import pdfplumber
        except ImportError as error:
            raise RuntimeError("install pdfplumber to extract PDF tables") from error
        tables: list[list[list[str]]] = []
        with pdfplumber.open(source) as document:
            for page in document.pages:
                for table in page.extract_tables():
                    tables.append(normalize_rows(table))
        return tables
    raise RuntimeError(f"unsupported table source: {extension or 'extensionless input'}")


def unique_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        base = value.strip() or f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def as_markdown(table: list[list[str]]) -> str:
    if not table:
        return ""
    width = max(len(row) for row in table)
    rows = [row + [""] * (width - len(row)) for row in table]

    def clean(value: str) -> str:
        return value.replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>")

    lines = [
        "| " + " | ".join(clean(value) for value in rows[0]) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    lines.extend("| " + " | ".join(clean(value) for value in row) + " |" for row in rows[1:])
    return "\n".join(lines) + "\n"


def write_atomic(destination: Path, data: str, encoding: str = "utf-8") -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding=encoding,
        newline="\n",
        delete=False,
        dir=destination.parent,
        prefix=f".{destination.name}.",
        suffix=".tmp",
    ) as handle:
        handle.write(data)
        temporary_path = Path(handle.name)
    try:
        os.replace(temporary_path, destination)
    finally:
        temporary_path.unlink(missing_ok=True)


def write_table(table: list[list[str]], destination: Path, output_format: str) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if output_format == "md":
        write_atomic(destination, as_markdown(table))
        return
    if output_format == "json":
        if not table:
            records = []
        else:
            headers = unique_headers(table[0])
            records = [
                dict(zip(headers, row + [""] * (len(headers) - len(row))))
                for row in table[1:]
            ]
        write_atomic(destination, json.dumps(records, ensure_ascii=False, indent=2) + "\n")
        return
    if output_format == "csv":
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8-sig",
            newline="",
            delete=False,
            dir=destination.parent,
            prefix=f".{destination.name}.",
            suffix=".tmp",
        ) as handle:
            csv.writer(handle).writerows(table)
            temporary_path = Path(handle.name)
        try:
            os.replace(temporary_path, destination)
        finally:
            temporary_path.unlink(missing_ok=True)
        return
    raise RuntimeError("output format must be csv, json, or md")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--sheet", help="XLSX worksheet name")
    parser.add_argument("--table", type=int, default=1, help="One-based table index")
    parser.add_argument("--format", choices=["csv", "json", "md"])
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    try:
        source = require_file(args.source)
        output = args.output.expanduser().resolve()
        if source == output:
            raise RuntimeError("source and output must differ")
        if output.exists() and not args.overwrite:
            raise FileExistsError(f"output already exists: {output}")
        tables = [table for table in load_tables(source, args.sheet) if table]
        if not tables:
            raise RuntimeError("no tables found")
        index = args.table - 1
        if index < 0 or index >= len(tables):
            raise RuntimeError(f"--table must be between 1 and {len(tables)}")
        output_format = args.format or output.suffix.lower().lstrip(".")
        if output_format == "markdown":
            output_format = "md"
        write_table(tables[index], output, output_format)
        print(f"Wrote {output}; discovered {len(tables)} table(s)")
        return 0
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
