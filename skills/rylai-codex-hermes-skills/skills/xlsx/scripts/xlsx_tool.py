#!/usr/bin/env python3
"""Original Rylai helper for practical XLSX workflows."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install it only with the user's approval."
    ) from exc


AUTHOR = "Rylai"
FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def set_workbook_author(workbook, author: str) -> None:
    workbook.properties.creator = author
    workbook.properties.lastModifiedBy = author


def safe_sheet_name(name: str) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("-", name).strip("'")[:31]
    return cleaned or "Data"


def coerce_value(value: str):
    text = value.strip()
    if text == "":
        return None
    if text.startswith("="):
        return tex
    if re.fullmatch(r"[+-]?\d+", text):
        unsigned = text.lstrip("+-")
        if len(unsigned) == 1 or not unsigned.startswith("0"):
            return int(text)
    if re.fullmatch(r"[+-]?(?:\d+\.\d*|\d*\.\d+)", text):
        return float(text)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass
    return value


def detect_delimiter(path: Path, requested: str) -> str:
    if requested == "comma":
        return ","
    if requested == "tab":
        return "\t"
    if requested == "semicolon":
        return ";"
    sample = path.read_text(encoding="utf-8-sig")[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
    except csv.Error:
        return "\t" if path.suffix.lower() == ".tsv" else ","


def create_workbook(
    source_path: Path,
    output_path: Path,
    sheet_name: str,
    delimiter_name: str,
    author: str,
) -> None:
    delimiter = detect_delimiter(source_path, delimiter_name)
    with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_sheet_name(sheet_name)
    set_workbook_author(workbook, author)
    workbook.properties.title = source_path.stem

    for row in rows:
        worksheet.append([coerce_value(value) for value in row])

    if rows:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    for column_index in range(1, worksheet.max_column + 1):
        values = [
            "" if worksheet.cell(row_index, column_index).value is None
            else str(worksheet.cell(row_index, column_index).value)
            for row_index in range(1, min(worksheet.max_row, 200) + 1)
        ]
        width = min(max((len(value) for value in values), default=8) + 2, 50)
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 10)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    load_workbook(output_path, read_only=True).close()


def inspect_workbook(path: Path) -> dict:
    workbook = load_workbook(path, data_only=False, read_only=False)
    formulas = []
    errors = []
    sheets = []

    for worksheet in workbook.worksheets:
        sheet_formulas = 0
        sheet_errors = 0
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append(f"{worksheet.title}!{cell.coordinate}")
                    sheet_formulas += 1
                if isinstance(value, str) and value.upper() in FORMULA_ERRORS:
                    errors.append(
                        {
                            "cell": f"{worksheet.title}!{cell.coordinate}",
                            "value": value,
                        }
                    )
                    sheet_errors += 1
        sheets.append(
            {
                "name": worksheet.title,
                "rows": worksheet.max_row,
                "columns": worksheet.max_column,
                "formulas": sheet_formulas,
                "errors": sheet_errors,
                "merged_ranges": len(worksheet.merged_cells.ranges),
            }
        )

    result = {
        "path": str(path.resolve()),
        "creator": workbook.properties.creator or "",
        "last_modified_by": workbook.properties.lastModifiedBy or "",
        "sheets": sheets,
        "formula_count": len(formulas),
        "formula_cells": formulas[:100],
        "error_count": len(errors),
        "errors": errors[:100],
    }
    workbook.close()
    return resul


def parse_cell_update(value: str) -> tuple[str, str, object]:
    if "=" not in value:
        raise SystemExit(f"Cell update must use Sheet!A1=value: {value}")
    target, raw_value = value.split("=", 1)
    if "!" not in target:
        raise SystemExit(f"Cell update must include a sheet name: {value}")
    sheet_name, coordinate = target.rsplit("!", 1)
    if not sheet_name or not re.fullmatch(r"\$?[A-Za-z]{1,3}\$?[1-9]\d*", coordinate):
        raise SystemExit(f"Invalid cell target: {target}")
    return sheet_name, coordinate.replace("$", "").upper(), coerce_value(raw_value)


def set_cells(
    input_path: Path,
    output_path: Path,
    updates: list[str],
    author: str,
) -> None:
    workbook = load_workbook(input_path)
    for update in updates:
        sheet_name, coordinate, value = parse_cell_update(update)
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        worksheet[coordinate] = value
    set_workbook_author(workbook, author)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    load_workbook(output_path, read_only=True).close()


def find_soffice() -> str | None:
    found = shutil.which("soffice") or shutil.which("libreoffice")
    if found:
        return found
    candidates = (
        Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
    )
    return next((str(path) for path in candidates if path.exists()), None)


def run_libreoffice(input_path: Path, output_dir: Path, target_format: str) -> Path:
    soffice = find_soffice()
    if not soffice:
        raise SystemExit(
            "LibreOffice was not found; XLSX creation, editing, and inspection still work."
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    command = [
        soffice,
        "--headless",
        "--convert-to",
        target_format,
        "--outdir",
        str(output_dir),
        str(input_path),
    ]
    completed = subprocess.run(command, capture_output=True, text=True, timeout=180)
    expected = output_dir / f"{input_path.stem}.{target_format}"
    if completed.returncode != 0 or not expected.exists():
        detail = completed.stderr.strip() or completed.stdout.strip() or "unknown error"
        raise SystemExit(f"LibreOffice conversion failed: {detail}")
    return expected


def recalculate(input_path: Path, output_path: Path) -> None:
    with tempfile.TemporaryDirectory(prefix="rylai-xlsx-") as temp_root:
        temp = Path(temp_root)
        source_dir = temp / "source"
        result_dir = temp / "result"
        source_dir.mkdir()
        result_dir.mkdir()
        staged = source_dir / input_path.name
        shutil.copy2(input_path, staged)
        recalculated = run_libreoffice(staged, result_dir, "xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(recalculated, output_path)
    load_workbook(output_path, read_only=True).close()


def render_workbook(input_path: Path, output_dir: Path) -> Path:
    return run_libreoffice(input_path, output_dir, "pdf")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Rylai XLSX helper")
    subparsers = parser.add_subparsers(dest="command", required=True)

    create = subparsers.add_parser("create", help="Create XLSX from CSV or TSV")
    create.add_argument("input", type=Path)
    create.add_argument("output", type=Path)
    create.add_argument("--sheet", default="Data")
    create.add_argument(
        "--delimiter",
        choices=("auto", "comma", "tab", "semicolon"),
        default="auto",
    )
    create.add_argument("--author", default=AUTHOR)

    inspect = subparsers.add_parser("inspect", help="Inspect workbook structure")
    inspect.add_argument("input", type=Path)

    set_command = subparsers.add_parser("set", help="Update cells in a workbook copy")
    set_command.add_argument("input", type=Path)
    set_command.add_argument("output", type=Path)
    set_command.add_argument("--cell", action="append", required=True)
    set_command.add_argument("--author", default=AUTHOR)

    recalc = subparsers.add_parser("recalc", help="Recalculate with LibreOffice")
    recalc.add_argument("input", type=Path)
    recalc.add_argument("output", type=Path)

    render = subparsers.add_parser("render", help="Render workbook to PDF")
    render.add_argument("input", type=Path)
    render.add_argument("output_dir", type=Path)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if args.command == "create":
        create_workbook(args.input, args.output, args.sheet, args.delimiter, args.author)
        print(json.dumps(inspect_workbook(args.output), ensure_ascii=False, indent=2))
    elif args.command == "inspect":
        print(json.dumps(inspect_workbook(args.input), ensure_ascii=False, indent=2))
    elif args.command == "set":
        set_cells(args.input, args.output, args.cell, args.author)
        print(json.dumps(inspect_workbook(args.output), ensure_ascii=False, indent=2))
    elif args.command == "recalc":
        recalculate(args.input, args.output)
        print(json.dumps(inspect_workbook(args.output), ensure_ascii=False, indent=2))
    else:
        print(render_workbook(args.input, args.output_dir))
    return 0


if __name__ == "__main__":
    sys.exit(main())
